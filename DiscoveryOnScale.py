import re
import logging
from napalm import get_network_driver
import pandas as pd
from openpyxl.styles import Font, PatternFill

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(_name_)

PLATFORMS_TO_OMIT_SSH = ['Cisco IP Phone 7841', 'Cisco IP Phone 7821', 'Cisco IP Phone 7942', 'Board Pro 55', 'Room 55', 'CTS-CODEC-DX80', 'cisco AIR-AP3802I-A-K9', 'cisco AIR-AP2802I-A-K9', 'cisco AIR-AP3802E-A-K9', 'cisco AIR-CAP2702I-A-K9', 'Room 70D', 'CTS-CODEC-MX200 G2', 'CTS-CODEC-inTouch', 'Desk Pro', 'cisco C9130AXI-A', 'Board Pro eth0', 'Board Pro 75', 'IW-6300H-DC-A-K9', 'cisco IW-6300H-DC-A-K9', 'Cisco-VM-SPID', 'CTS-CODEC-MX300 G2', 'ISE-VM-K9', 'MikroTik', 'Room 55', 'Room 70S G2', 'Room Bar', 'Room Kit Mini', 'cisco AIR-AP3802I-N-K9', 'cisco AIR-AP1852E-A-K9', 'cisco AIR-AP3802I-H-K9', 'cisco AIR-AP1562I-A-K9']

def get_device_type(platform):
    if any(platform.startswith(p) for p in ['cisco C9300-48UXM', 'cisco C9200L-24P-4G', 'cisco C9200-48P', 'cisco C9300-24T', 'cisco WS-C3850-48P', 'cisco IE-3400-8P2S', 'cisco C9500-16X', 'cisco C8300-2N2S-4T2X', 'cisco WS-C2960X-48FPS-L', 'cisco WS-C3850-24P', 'cisco WS-C2960X-48FPS-L', 'cisco WS-C2960X-48FPD-L', 'cisco C9500-48Y4C', 'cisco C9500-40X', 'cisco C9300L-24P-4G', 'cisco ISR4351/K9', 'cisco C9300-48P', 'cisco C9200-24PXG', 'cisco C9200L-24P-4X', 'cisco C9300-24P', 'cisco C9300-48P', 'cisco C9300L-24P-4X', 'cisco C9300L-24T-4G', 'cisco C9300L-48P-4G', 'cisco IE-2000-8TC-G-L', 'cisco IE-4010-16S12P', 'cisco WS-C2960-24PC-L', 'cisco WS-C2960-48PST-L', 'cisco WS-C2960X-24PS-L', 'cisco WS-C3560CX-12PC-S', 'cisco WS-C3560CX-8PC-S', 'cisco C9200-24P', 'cisco IE-3400-8T2S', 'cisco C9200L-24T-4G', 'cisco IE-4000-8T4G-E', 'cisco IE-2000-16TC-G-L']): return 'Switch'
    elif any(platform.startswith(p) for p in ['cisco AIR-AP2802I-A-K9', 'cisco C9130AXI-A', 'cisco AIR-AP3802I-A-K9', 'IW-6300H-DC-A-K9', 'cisco AIR-AP1852E-A-K9', 'cisco AIR-AP3802I-H-K9', 'cisco AIR-AP3802I-N-K9', 'cisco IW-6300H-DC-A-K9', 'MikroTik', 'cisco AIR-AP1562I-A-K9']): return 'AP'
    elif any(platform.startswith(p) for p in ['Cisco IP Phone 7821', 'Cisco IP Phone 7841', 'CTS-CODEC-DX80', 'CTS-CODEC-MX300 G2']): return 'IP Phone'
    elif any(platform.startswith(p) for p in ['Board Pro 75', 'Board Pro 55', 'Desk Pro', 'Room 70D', 'Room 55', 'Room 70S G2', 'Room Bar', 'Room Kit Mini']): return 'VC'
    else: return 'N/A'

def should_skip_ssh(platform, neighbor_ip):
    if any(omit_platform.lower() in platform.lower() for omit_platform in PLATFORMS_TO_OMIT_SSH):
        logger.info(f"Skipping SSH connection to device with platform: {platform} and IP: {neighbor_ip}")
        return True
    elif neighbor_ip in ['10.10.1.1', '10.4.206.133', '10.4.206.134', '10.180.20.30', '10.180.20.14', '10.179.237.10', '10.179.235.147', '10.179.235.146', 'N/A', '10.179.4.230', '10.184.134.201', '10.179.21.100', '10.179.239.141']:
        logger.info(f"Skipping SSH connection to device with IP: {neighbor_ip}")
        return True
    else:
        return False

def get_cdp_neighbors(device_instance, device):
    try:
        platform = device.get('platform', '').strip()
        if should_skip_ssh(platform, device['ip']):
            return [{'platform': platform, 'status': 'Omitido'}]
        facts = device_instance.get_facts()
        hostname = facts.get('hostname', 'N/A')
        cdp_output = device_instance.cli(["show cdp neighbors detail"])
        return parse_cdp_neighbors(cdp_output['show cdp neighbors detail'], hostname)
    except Exception as e:
        logger.error(f"Error processing device {device['ip']}: {e}")
        if "TCP connection to device failed" in str(e):
            logger.warning(f"Skipping device {device['ip']} due to TCP connection failure. {str(e)}")
            return []
        elif "getaddrinfo failed" in str(e):
            logger.warning(f"Skipping device {device['ip']} due to SSH connection failure. {str(e)}")
            return []
        else:
            logger.warning(f"Skipping device {device['ip']} due to unknown error. {str(e)}")
            return []

def parse_cdp_neighbors(cdp_output, hostname):
    cdp_neighbors = []
    current_neighbor = {}
    for line in cdp_output.splitlines():
        if 'Device ID: ' in line:
            if current_neighbor: cdp_neighbors.append(current_neighbor)
            current_neighbor = {'neighbor': line.split('Device ID: ')[1].strip(), 'hostname': hostname}
        elif 'IP address: ' in line:
            current_neighbor['ip_address'] = line.split('IP address: ')[1].strip()
        elif 'Interface: ' in line:
            current_neighbor['local_interface'] = line.split('Interface: ')[1].split(',')[0].strip()
            current_neighbor['neighbor_interface'] = line.split('Port ID (outgoing port): ')[1].strip()
        elif 'Platform: ' in line:
            current_neighbor['platform'] = line.split('Platform: ')[1].split(',')[0].strip()
    if current_neighbor: cdp_neighbors.append(current_neighbor)
    return cdp_neighbors

def parse_version_output(version_output):
    version_data = {}
    mac_address_match = re.findall(r"Base ethernet MAC Address\s*:\s*(\S+)", version_output, re.IGNORECASE)
    version_data['MAC Address'] = ', '.join(mac_address_match) if mac_address_match else 'N/A'

    serial_number_match = re.findall(r"System serial number\s*:\s*(\S+)", version_output, re.IGNORECASE)
    version_data['Serial Number'] = ', '.join(serial_number_match) if serial_number_match else 'N/A'

    software_version_match = re.search(r"\\s+(.)", version_output, re.IGNORECASE)
    if software_version_match:
        software_version = software_version_match.group(1).strip().split()
        version_data['Model'] = software_version[2] if len(software_version) > 2 else 'N/A'
        version_data['Firmware version'] = software_version[3] if len(software_version) > 3 else 'N/A'
        version_data['SW Image'] = software_version[4] if len(software_version) > 4 else 'N/A'
    else:
        version_data['Model'], version_data['Firmware version'], version_data['SW Image'] = 'N/A', 'N/A', 'N/A'

    # Buscar palabras específicas en la segunda línea
    lines = version_output.split('\n')
    if len(lines) > 1:
        second_line_words = lines[1].split()
        if len(second_line_words) > 7:
            version_data['SW Image'] = second_line_words[7]
            if len(second_line_words) > 9:
                version_data['Firmware version'] = second_line_words[9]

    return version_data

def main():
    devices = [{'ip': '10.180.241.129', 'username': 'xhcldarroyo', 'password': 'DerthGamer123??', 'secret': 'DerthGamer123??'},{'ip': '10.184.135.193', 'username': 'xhcldarroyo', 'password': 'DerthGamer123??', 'secret': 'DerthGamer123??'}, {'ip': '10.179.4.1', 'username': 'xhcldarroyo', 'password': 'DerthGamer123??', 'secret': 'DerthGamer123??'}, {'ip': '10.179.8.2', 'username': 'xhcldarroyo', 'password': 'DerthGamer123??', 'secret': 'DerthGamer123??'}, {'ip': '10.179.235.137', 'username': 'xhcldarroyo', 'password': 'DerthGamer123??', 'secret': 'DerthGamer123??'}, {'ip': '10.179.235.136', 'username': 'xhcldarroyo', 'password': 'DerthGamer123??', 'secret': 'DerthGamer123??'}, {'ip': '10.179.235.132', 'username': 'xhcldarroyo', 'password': 'DerthGamer123??', 'secret': 'DerthGamer123??'}, {'ip': '10.179.235.129', 'username': 'xhcldarroyo', 'password': 'DerthGamer123??', 'secret': 'DerthGamer123??'}, {'ip': '10.179.251.1', 'username': 'xhcldarroyo', 'password': 'DerthGamer123??', 'secret': 'DerthGamer123??'}, {'ip': '10.179.25.78', 'username': 'xhcldarroyo', 'password': 'DerthGamer123??', 'secret': 'DerthGamer123??'}, {'ip': '10.179.240.65', 'username': 'xhcldarroyo', 'password': 'DerthGamer123??', 'secret': 'DerthGamer123??'}, {'ip': '10.179.234.129', 'username': 'xhcldarroyo', 'password': 'DerthGamer123??', 'secret': 'DerthGamer123??'}]
    processed_ips, table_data, sw_info_data, stack_data = set(), [], [], []
    headers = ["Device Name", "Device IP", "Neighbor's Device Name", "Neighbor's Device IP", "Local Interface", "Neighbor's Interface", "Neighbor's Platform", "Neighbor's Category", "Status", "MAC Address", "Serial Number", "Model", "Firmware version", "SW Image"]
    sw_info_headers = ["Device Name", "Device IP", "MAC Address", "Serial Number", "Model", "Firmware version", "SW Image"]

    while devices:
        current_device = devices.pop(0)
        if current_device['ip'] in processed_ips: continue
        used_ips = set()

        driver = get_network_driver('ios')
        try:
            device_instance = driver(hostname=current_device['ip'], username=current_device['username'], password=current_device['password'])
            device_instance.open()
        except Exception as e:
            logger.error(f"Failed to connect to device {current_device['ip']} with primary credentials: {e}")
            logger.info(f"Attempting to connect with alternative credentials...")
            try:
                device_instance = driver(hostname=current_device['ip'], username='hcladmin_aa', password='fr33d0m', optional_args={'secret': 'fr33d0m'})
                device_instance.open()
            except Exception as e:
                logger.error(f"Failed to connect to device {current_device['ip']} with alternative credentials: {e}")
                continue

        cdp_neighbors = get_cdp_neighbors(device_instance, current_device)
        if not cdp_neighbors:
            logger.info(f"No neighbors found for {current_device['ip']}")
            processed_ips.add(current_device['ip'])
            device_instance.close()
            continue

        for neighbor in cdp_neighbors:
            neighbor_ip = neighbor.get('ip_address', 'N/A')
            neighbor_platform = neighbor.get('platform', '').strip()

            if should_skip_ssh(neighbor_platform, neighbor_ip):
                status = 'Omitido'
            else:
                status = 'Conectado'
                if neighbor_ip not in used_ips:
                    used_ips.add(neighbor_ip)
                    devices.append({'ip': neighbor_ip, 'username': current_device['username'], 'password': current_device['password'], 'secret': current_device['secret']})

            try:
                version_data = parse_version_output(device_instance.cli(["show version | begin Base"])['show version | begin Base'])
            except Exception as e:
                logger.error(f"Error retrieving version information for {current_device['ip']}: {e}")
                version_data = {'MAC Address': 'N/A', 'Serial Number': 'N/A', 'Model': 'N/A', 'Firmware version': 'N/A', 'SW Image': 'N/A'}

            table_data.append([neighbor.get('hostname', 'N/A'), current_device['ip'], neighbor.get('neighbor', 'N/A'), neighbor_ip, neighbor.get('local_interface', 'N/A'), neighbor.get('neighbor_interface', 'N/A'), neighbor_platform, get_device_type(neighbor_platform), status, version_data['MAC Address'], version_data['Serial Number'], version_data['Model'], version_data['Firmware version'], version_data['SW Image']])
            sw_info_data.append([current_device['ip'], neighbor.get('hostname', 'N/A'), version_data['MAC Address'], version_data['Serial Number'], version_data['Model'], version_data['Firmware version'], version_data['SW Image']])
            if ',' in version_data['MAC Address'] or ',' in version_data['Serial Number']:
                stack_data.append([neighbor.get('hostname', 'N/A'), current_device['ip'], version_data['MAC Address'], version_data['Serial Number'], version_data['Model'], version_data['Firmware version'], version_data['SW Image']])
            logger.info(f"Recopilando información de vecino en {neighbor_ip} (Modelo: {version_data['Model']})...")

        processed_ips.add(current_device['ip'])
        device_instance.close()

    df = pd.DataFrame(table_data, columns=headers)
    print(df.to_string(index=False))

    sw_info_data_adjusted = [[row[1], row[0], *row[2:]] for row in sw_info_data]
    df_sw_info = pd.DataFrame(sw_info_data_adjusted, columns=sw_info_headers)

    df_sw_info.drop_duplicates(subset=["Device Name", "Device IP"], keep="first", inplace=True)
    df_stack_info = pd.DataFrame(stack_data, columns=sw_info_headers)
    df_stack_info.drop_duplicates(subset=["Device Name", "Device IP"], keep="first", inplace=True)

    df_sw_final = pd.DataFrame(sw_info_data_adjusted, columns=sw_info_headers)
    df_sw_final.drop_duplicates(subset=["Device Name", "Device IP"], keep="first", inplace=True)

    save_to_excel = input("¿Desea guardar la salida en un archivo Excel? (yes/no): ").lower() == 'yes'
    if save_to_excel:
        file_name = input("Ingrese el nombre del archivo Excel (sin extensión): ") or 'output'
        output_file_path = rf'C:\Users\xhcldarrollo\OneDrive - Anglo American\Desktop\{file_name}.xlsx'
        with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Full_Network_Data', index=False)
            apply_formatting_to_sheet(writer.sheets['Full_Network_Data'])
            df_sw_info.to_excel(writer, sheet_name='List_SW_Information', index=False)
            apply_formatting_to_sheet(writer.sheets['List_SW_Information'])
            df_sw_final.to_excel(writer, sheet_name='List_Switches_Final', index=False)
            create_list_switches_final(writer, sw_info_data_adjusted)
            create_list_ap(writer, df)
            create_list_VC(writer, df)
            create_list_Phone(writer, df)
        logger.info(f"Output guardado en: {output_file_path}")



def apply_formatting_to_sheet(ws):
    # Establecer el formato de las cabeceras
    for cell in ws[1]:
        cell.font = Font(name='Aptos Narrow', size=12, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="0E82E2", fill_type="solid")

    # Establecer el ancho de las columnas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Obtener la letra de la columna
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width

def create_list_switches_final(writer, sw_info_data_adjusted):
    df_final = pd.DataFrame(columns=["Device Name", "Device IP", "MAC Address", "Serial Number", "Model", "Firmware version", "SW Image"])
    for row in sw_info_data_adjusted:
        device_name, device_ip, mac_address, serial_number, model, firmware_version, sw_image = row
        if ',' in mac_address:
            mac_addresses = mac_address.split(', ')
            serial_numbers = serial_number.split(', ')
            for idx, (mac, serial) in enumerate(zip(mac_addresses, serial_numbers), start=1):
                df_final = pd.concat([df_final, pd.DataFrame([[f"{device_name}-{idx}", device_ip, mac, serial, model, firmware_version, sw_image]], columns=df_final.columns)], ignore_index=True)
        else:
            df_final = pd.concat([df_final, pd.DataFrame([row], columns=df_final.columns)], ignore_index=True)
    df_final.drop_duplicates(subset=["Device Name", "Device IP"], keep="first", inplace=True)
    df_final.to_excel(writer, sheet_name='List_Switches_Final', index=False)

    # Obtener la hoja de Excel
    ws_combined = writer.sheets['List_Switches_Final']
    # Aplicar formato a la hoja
    apply_formatting_to_sheet(ws_combined)

def create_list_ap(writer, df):
    df_ap = df[df["Neighbor's Category"] == "AP"]
    selected_columns = ["Neighbor's Device Name", "Neighbor's Device IP", "Neighbor's Platform","Device Name", "Local Interface", "Device IP"]
    df_ap_selected = df_ap[selected_columns]
    df_ap_selected.columns = ["Device Name AP", "IP AP", "Model", "Uplink Switch [Network Gear]", "Uplink Switch Port [Network Gear]", "IP Uplink Switch"]
    df_ap_selected.to_excel(writer, sheet_name='List_AP', index=False)
    ws_ap = writer.sheets['List_AP']
    apply_formatting_to_sheet(ws_ap)

def create_list_VC(writer, df):
    df_vc = df[df["Neighbor's Category"] == "VC"]
    selected_columns = ["Neighbor's Device Name", "Neighbor's Device IP", "Neighbor's Platform", "Device Name", "Local Interface", "Device IP"]
    df_vc_selected = df_vc[selected_columns]
    df_vc_selected.columns = ["Device Name VC", "IP VC", "Model", "Uplink Switch [Network Gear]", "Uplink Switch Port [Network Gear]", "IP Uplink Switch"]
    df_vc_selected.to_excel(writer, sheet_name='List_VC', index=False)
    ws_vc = writer.sheets['List_VC']
    apply_formatting_to_sheet(ws_vc)

def create_list_Phone(writer, df):
    df_phone = df[df["Neighbor's Category"] == "IP Phone"]
    selected_columns = ["Neighbor's Device Name", "Neighbor's Device IP", "Neighbor's Platform", "Device Name", "Local Interface", "Device IP"]
    df_phone_selected = df_phone[selected_columns]
    df_phone_selected.columns = ["Device Name Phone", "IP Phone", "Model", "Uplink Switch [Network Gear]", "Uplink Switch Port [Network Gear]", "IP Uplink Switch"]
    df_phone_selected.to_excel(writer, sheet_name='List_Phone', index=False)
    ws_phone = writer.sheets['List_Phone']
    apply_formatting_to_sheet(ws_phone)

if _name_ == "_main_":
    main()
